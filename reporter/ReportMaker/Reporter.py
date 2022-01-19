import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os


class Reporter:

    def __init__(self):
        self.wb = Workbook()  # openpyxl workbook
        self.worksheets = []  # List of openpyxl worksheets

    def write(self, file_path, filename):
        ws = self.wb.create_sheet(title=filename.split(".csv")[0])  # Get only the filename without .csv extension

        with open(file_path, "r") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)

        self.worksheets.append(ws)

    def save(self, path):
        self.align()

        if not path.endswith(".xlsx"):
            path += ".xlsx"

        self.wb.save(path)

    ###########################
    #    APPEARANCE METHODS   #
    ##########################

    def align(self):
        """Makes necessary actions to make the excel presentable
        EX: column width, data alignment"""

        for ws in self.worksheets:
            self.align_data_to_center(ws)
            self.adjust_column_width(ws)

    def align_data_to_center(self, ws):
        alignment = Alignment(horizontal="center", vertical="center")

        # Align data to center
        for col in ws.columns:
            for cell in col:
                cell.alignment = alignment

    def adjust_column_width(self, ws):
        # Column Width
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 1.5  # +1.5 to add a bit more space for each column
